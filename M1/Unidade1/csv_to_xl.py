import csv
import openpyxl

xl_file = '/home/boris/PycharmProjects/TokioSchoolML/M1/Unidade1/solutions.xlsx'
wb = openpyxl.load_workbook(xl_file)

file_lists = ['/home/boris/PycharmProjects/TokioSchoolML/M1/Unidade1/M1_U1_Ej1_dataset-tarea1.csv',
              '/home/boris/PycharmProjects/TokioSchoolML/M1/Unidade1/M1_U1_Ej1_dataset-tarea2.csv',
              '/home/boris/PycharmProjects/TokioSchoolML/M1/Unidade1/M1_U1_Ej1_dataset-tarea3.csv',
              '/home/boris/PycharmProjects/TokioSchoolML/M1/Unidade1/M1_U1_Ej1_dataset-tarea4.csv'
              ]

for obj in file_lists:
    sn = obj.split('.csv')
    sheet_name = sn[0][-6:]
    with open(obj) as file_path:
        file = csv.reader(file_path, delimiter=' ', quotechar='|')
        if 'Sheet1' in wb.sheetnames:
            wb['Sheet1'].title = sheet_name
        elif sheet_name not in wb.sheetnames:
            wb.create_sheet(title=sheet_name)
        for i, row in enumerate(file):
            if i < 1:
                res = row[0].split(",")
            else:
                res = row[0][:-1].split(',"')
                if ',' in res[1]:
                    res[1] = res[1].replace(',', '.')
            # print(res, res[1])
            cellA = 'A' + str(i + 1)
            cellB = 'B' + str(i + 1)
            try:
                wb[sheet_name][cellA].value = int(res[0])
                wb[sheet_name][cellB].value = float(res[1])
                # print(wb[sheet_name][cellA].value, type(wb[sheet_name][cellA].value), wb[sheet_name][cellB].value,
                #       type(wb[sheet_name][cellB].value))
            except:
                pass

wb.save(xl_file)
