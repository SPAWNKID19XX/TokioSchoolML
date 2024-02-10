import csv
import openpyxl

xl_file = 'M1/solutions.xlsx'
wb = openpyxl.load_workbook(xl_file)


file_lists= ['M1/M1_U1_Ej1_dataset-tarea1.csv', 
             'M1/M1_U1_Ej1_dataset-tarea2.csv', 
             'M1/M1_U1_Ej1_dataset-tarea3.csv',
             'M1/M1_U1_Ej1_dataset-tarea4.csv'
             ]

for obj in file_lists:
    sn = obj.split('.csv')
    sheet_name = sn[0][-6:]
    with open(obj) as file_path:
        file=csv.reader(file_path, delimiter=' ',quotechar='|')
        if 'Sheet1' in wb.sheetnames:
            wb['Sheet1'].title = sheet_name
        elif sheet_name not in wb.sheetnames:
            wb.create_sheet(title=sheet_name)
        for i, row in enumerate(file):
            if i < 1:    
                res = row[0].split(",")
            else:
                res = row[0][:-1].split(',"')
                if ',' in res[1]:
                    res[1] = res[1].replace(',' , '.')
            print(res, res[1])
            cellA = 'A' + str(i+1)
            cellB = 'B' + str(i+1)
            wb[sheet_name][cellA].value = res[0]
            wb[sheet_name][cellB].value = res[1]
        

wb.save(xl_file)
